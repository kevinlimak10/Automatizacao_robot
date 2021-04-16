from openpyxl import load_workbook
from pathlib import Path
from robot.api.deco import keyword

@keyword
def obter_dados_excel():

    pasta_existente = load_workbook('./excel/tabela_calculo.xlsx')

    planilhasNomes = pasta_existente.sheetnames

    print(planilhasNomes)

    planilha = pasta_existente.worksheets[0]

    List = []

    for linha in planilha.values:
        for valor in linha:
            if valor == None: valor =''
            List.append(valor)

    print(List)
    return List

